"""
Lector del Plan de Producción Mensual directo del Excel -- sin depender de las
tablas programa_produccion_* en ia_fasa (bloqueadas: el usuario `sapiens` es
solo-lectura, ver scripts/ddl_programa_produccion.sql). Mismo parser que
scripts/cargar_programa_produccion.py, pero de solo lectura -- nunca escribe
nada, ni a la BD ni al archivo.

Cuando IT/Jasso otorguen permiso de escritura, /api/programa puede empezar a
preferir la BD (más rápido, histórico de revisiones) y usar esto como fallback.
"""
import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

PLAN_DIR = Path(r"C:\Users\Administrator\Documents\FASA\plan de produccion mensual")

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
MESES_NOMBRE = {v: k.capitalize() for k, v in MESES.items()}


def _parse_anio_mes(titulo: str) -> Optional[date]:
    m = re.search(r"([A-ZÑ]+)\s+(\d{4})", (titulo or "").upper())
    if not m or m.group(1) not in MESES:
        return None
    return date(int(m.group(2)), MESES[m.group(1)], 1)


def _archivo_para_mes(mes: str) -> Optional[Path]:
    """mes='2026-08' -> busca un archivo 'Plan 08*Agosto*.xlsx' en PLAN_DIR.
    El nombre de archivo (ej. 'Plan 08 Agosto.xlsx') no trae año -- la
    validación real del año pasa después, comparando contra el título dentro
    del Excel (_parse_anio_mes), no contra el nombre de archivo."""
    if not PLAN_DIR.is_dir():
        return None
    try:
        _, mm = mes.split("-")
        mes_num = int(mm)
    except ValueError:
        return None
    nombre_mes = MESES_NOMBRE.get(mes_num)
    if not nombre_mes:
        return None
    candidatos = [
        p for p in PLAN_DIR.glob(f"Plan {mm}*.xlsx")
        if not p.name.startswith("~$") and nombre_mes.upper() in p.name.upper()
    ]
    return candidatos[0] if candidatos else None


def leer_plan(mes: str) -> Optional[dict]:
    """Devuelve el mismo shape que /api/programa (encabezado, ritmos_por_proceso,
    ritmos_agregados, detalle_referencial), leído en vivo del Excel. None si no
    hay archivo para ese mes, o si el título dentro del Excel no coincide con
    el mes/año pedido (protección contra el archivo mal nombrado/año equivocado)."""
    archivo = _archivo_para_mes(mes)
    if not archivo:
        return None

    wb = openpyxl.load_workbook(archivo, data_only=True)
    if "Plan" not in wb.sheetnames or "resumen" not in wb.sheetnames:
        return None
    plan_ws = wb["Plan"]
    resumen_ws = wb["resumen"]

    anio_mes = _parse_anio_mes(plan_ws["A1"].value)
    if not anio_mes or anio_mes.strftime("%Y-%m") != mes:
        return None  # el archivo encontrado no es el mes/año que se pidió

    encabezado = {
        "anio_mes":       str(anio_mes),
        "rev_no":         plan_ws["F4"].value,
        "fecha_rev":      str(plan_ws["F5"].value) if plan_ws["F5"].value else None,
        "dias_habiles":   plan_ws["C4"].value,
        "ventas_ton":     plan_ws["B7"].value,
        "presup_ton":     plan_ws["C7"].value,
        "buenas_ton":     plan_ws["B8"].value,
        "vaciadas_ton":   plan_ws["B9"].value,
        "rech_int_pct":   plan_ws["C9"].value,
        "linea_ton":      plan_ws["B16"].value,
        "desarrollo_ton": plan_ws["B17"].value,
        "linea_pct":      plan_ws["C16"].value,
        "desarrollo_pct": plan_ws["C17"].value,
        "inv_total":      plan_ws["F13"].value,
        "archivo_origen": archivo.name,
    }

    detalle_referencial = []
    for row in range(3, 141):  # filas 3-140 de `resumen`
        parte = resumen_ws.cell(row=row, column=2).value
        if parte in (None, ""):
            continue
        detalle_referencial.append({
            "parte":         str(parte),
            "sdo_final":     resumen_ws.cell(row=row, column=9).value,   # I
            "peso_kg":       resumen_ws.cell(row=row, column=11).value,  # K
            "kg_buenos":     resumen_ws.cell(row=row, column=12).value,  # L
            "proceso":       resumen_ws.cell(row=row, column=15).value,  # O
            "status":        resumen_ws.cell(row=row, column=17).value,  # Q
            "autoritativo":  False,
        })

    ritmos_por_proceso = []
    for tipo_dia, rango in (("LV", range(21, 24)), ("SAB", range(30, 33))):
        for row in rango:
            ritmos_por_proceso.append({
                "tipo_dia":     tipo_dia,
                "proceso":      plan_ws.cell(row=row, column=1).value,
                "moldes_dia":   plan_ws.cell(row=row, column=2).value,
                "peso_prom_kg": plan_ws.cell(row=row, column=3).value,
                "kg_diarios":   plan_ws.cell(row=row, column=4).value,
            })

    ritmos_agregados = [
        {
            "tipo_dia": "LV",
            "kg_vaciado_dia_con_ri": plan_ws["D24"].value,
            "kg_diarios_buenos":     plan_ws["D25"].value,
            "coladas_diarias":       plan_ws["D26"].value,
        },
        {
            "tipo_dia": "SAB",
            "kg_vaciado_dia_con_ri": plan_ws["D33"].value,
            "kg_diarios_buenos":     plan_ws["D34"].value,
            "coladas_diarias":       plan_ws["D35"].value,
        },
    ]

    return {
        "fuente":              "excel_vivo",
        "encabezado":          encabezado,
        "ritmos_por_proceso":  ritmos_por_proceso,
        "ritmos_agregados":    ritmos_agregados,
        "detalle_referencial": detalle_referencial,
    }
