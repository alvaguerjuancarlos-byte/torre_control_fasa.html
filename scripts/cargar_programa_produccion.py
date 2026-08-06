"""
Carga del Plan de Produccion Mensual (Excel) a ia_fasa.
Tablas: programa_produccion_mensual, programa_produccion_detalle_referencial,
        programa_ritmos_vaciado, programa_ritmos_agregados.

NO EJECUTAR contra la BD hasta que el DDL (scripts/ddl_programa_produccion.sql)
haya sido corrido y confirmado. Con --dry-run solo lee el .xlsx y valida los
checksums estructurales del Paso 3, sin tocar la base de datos.

Checksum estructural (no de negocio):
  - 138 filas en programa_produccion_detalle_referencial (filas 3-140 de `resumen`)
  - 6 filas en programa_ritmos_vaciado (3 procesos AF1/AF2/AF3 x 2 tipos_dia)
  - 2 filas en programa_ritmos_agregados (1 fila por tipo_dia, esquema ancho
    con 3 metricas como columnas -- NO 6; ver decision del 2026-08-05)
No se reconcilia kg_buenos de la tabla 2 contra ningun total de la hoja Plan
(fila 142 de `resumen` no es un checksum valido, son fuentes distintas).

Uso:
  python cargar_programa_produccion.py --dry-run
  python cargar_programa_produccion.py            # carga real (requiere DDL ya aplicado)
"""
import argparse
import os
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

BASE = Path(__file__).resolve().parent.parent
load_dotenv(BASE / ".env")

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def parse_anio_mes(titulo: str) -> date:
    """'PLAN DE EMBARQUES Y PRODUCCION AGOSTO 2026' -> date(2026, 8, 1)"""
    m = re.search(r"([A-ZÑ]+)\s+(\d{4})", titulo.upper())
    if not m or m.group(1) not in MESES:
        raise ValueError(f"No se pudo derivar anio_mes del titulo: {titulo!r}")
    return date(int(m.group(2)), MESES[m.group(1)], 1)


def leer_encabezado(plan_ws, archivo_origen: str) -> dict:
    return {
        "anio_mes": parse_anio_mes(plan_ws["A1"].value),
        "rev_no": plan_ws["F4"].value,
        "fecha_rev": plan_ws["F5"].value,
        "dias_habiles": plan_ws["C4"].value,
        "ventas_ton": plan_ws["B7"].value,
        "presup_ton": plan_ws["C7"].value,
        "buenas_ton": plan_ws["B8"].value,
        "vaciadas_ton": plan_ws["B9"].value,
        "rech_int_pct": plan_ws["C9"].value,
        "linea_ton": plan_ws["B16"].value,
        "desarrollo_ton": plan_ws["B17"].value,
        "linea_pct": plan_ws["C16"].value,
        "desarrollo_pct": plan_ws["C17"].value,
        "inv_total": plan_ws["F13"].value,
        "archivo_origen": archivo_origen,
    }


def leer_detalle_referencial(resumen_ws) -> list[dict]:
    filas = []
    for row in range(3, 141):  # filas 3-140
        parte = resumen_ws.cell(row=row, column=2).value  # B
        if parte in (None, ""):
            continue
        filas.append({
            "parte": str(parte),
            "sdo_final": resumen_ws.cell(row=row, column=9).value,   # I
            "peso_kg": resumen_ws.cell(row=row, column=11).value,    # K
            "kg_buenos": resumen_ws.cell(row=row, column=12).value,  # L
            "proceso": resumen_ws.cell(row=row, column=15).value,    # O
            "status": resumen_ws.cell(row=row, column=17).value,     # Q
            "es_autoritativo": False,
        })
    return filas


def leer_ritmos_vaciado(plan_ws) -> list[dict]:
    filas = []
    bloques = [("LV", range(21, 24)), ("SAB", range(30, 33))]
    for tipo_dia, rango in bloques:
        for row in rango:
            filas.append({
                "tipo_dia": tipo_dia,
                "proceso": plan_ws.cell(row=row, column=1).value,       # A
                "moldes_dia": plan_ws.cell(row=row, column=2).value,    # B
                "peso_prom_kg": plan_ws.cell(row=row, column=3).value,  # C
                "kg_diarios": plan_ws.cell(row=row, column=4).value,    # D
            })
    return filas


def leer_ritmos_agregados(plan_ws) -> list[dict]:
    # una fila por tipo_dia, esquema ancho (3 metricas -> 3 columnas)
    return [
        {
            "tipo_dia": "LV",
            "kg_vaciado_dia_con_ri": plan_ws["D24"].value,
            "kg_diarios_buenos": plan_ws["D25"].value,
            "coladas_diarias": plan_ws["D26"].value,
        },
        {
            "tipo_dia": "SAB",
            "kg_vaciado_dia_con_ri": plan_ws["D33"].value,
            "kg_diarios_buenos": plan_ws["D34"].value,
            "coladas_diarias": plan_ws["D35"].value,
        },
    ]


def validar_checksums(detalle, vaciado, agregados):
    errores = []
    if len(detalle) != 138:
        errores.append(f"detalle_referencial: se esperaban 138 filas, se leyeron {len(detalle)}")
    if len(vaciado) != 6:
        errores.append(f"ritmos_vaciado: se esperaban 6 filas, se leyeron {len(vaciado)}")
    if len(agregados) != 2:
        errores.append(f"ritmos_agregados: se esperaban 2 filas, se leyeron {len(agregados)}")
    return errores


def cargar(xlsx_path: Path, dry_run: bool):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    plan_ws = wb["Plan"]
    resumen_ws = wb["resumen"]

    encabezado = leer_encabezado(plan_ws, xlsx_path.name)

    detalle = leer_detalle_referencial(resumen_ws)
    vaciado = leer_ritmos_vaciado(plan_ws)
    agregados = leer_ritmos_agregados(plan_ws)

    errores = validar_checksums(detalle, vaciado, agregados)
    print(f"Encabezado: anio_mes={encabezado['anio_mes']} rev_no={encabezado['rev_no']} "
          f"archivo_origen={encabezado['archivo_origen']}")
    print(f"Checksum -> detalle_referencial={len(detalle)} filas, "
          f"ritmos_vaciado={len(vaciado)} filas, ritmos_agregados={len(agregados)} filas")

    if errores:
        print("\nCHECKSUM FALLIDO:")
        for e in errores:
            print(f"  - {e}")
        sys.exit(1)
    print("Checksum OK.")

    if dry_run:
        print("\n--dry-run: no se escribio nada en la base de datos.")
        return

    HOST = os.environ["DB_HOST"]
    PORT = int(os.environ.get("DB_PORT", 3306))
    USER = os.environ["DB_USER"]
    PWD = os.environ["DB_PWD"]
    engine = create_engine(f"mysql+pymysql://{USER}:{PWD}@{HOST}:{PORT}/ia_fasa?charset=utf8mb4")

    with engine.begin() as conn:
        res = conn.execute(text("""
            INSERT INTO programa_produccion_mensual
                (anio_mes, rev_no, fecha_rev, dias_habiles, ventas_ton, presup_ton,
                 buenas_ton, vaciadas_ton, rech_int_pct, linea_ton, desarrollo_ton,
                 linea_pct, desarrollo_pct, inv_total, archivo_origen)
            VALUES
                (:anio_mes, :rev_no, :fecha_rev, :dias_habiles, :ventas_ton, :presup_ton,
                 :buenas_ton, :vaciadas_ton, :rech_int_pct, :linea_ton, :desarrollo_ton,
                 :linea_pct, :desarrollo_pct, :inv_total, :archivo_origen)
        """), encabezado)
        programa_id = res.lastrowid

        for fila in detalle:
            conn.execute(text("""
                INSERT INTO programa_produccion_detalle_referencial
                    (programa_id, parte, sdo_final, peso_kg, kg_buenos, proceso, status, es_autoritativo)
                VALUES
                    (:programa_id, :parte, :sdo_final, :peso_kg, :kg_buenos, :proceso, :status, :es_autoritativo)
            """), {**fila, "programa_id": programa_id})

        for fila in vaciado:
            conn.execute(text("""
                INSERT INTO programa_ritmos_vaciado
                    (programa_id, tipo_dia, proceso, moldes_dia, peso_prom_kg, kg_diarios)
                VALUES
                    (:programa_id, :tipo_dia, :proceso, :moldes_dia, :peso_prom_kg, :kg_diarios)
            """), {**fila, "programa_id": programa_id})

        for fila in agregados:
            conn.execute(text("""
                INSERT INTO programa_ritmos_agregados
                    (programa_id, tipo_dia, kg_vaciado_dia_con_ri, kg_diarios_buenos, coladas_diarias)
                VALUES
                    (:programa_id, :tipo_dia, :kg_vaciado_dia_con_ri, :kg_diarios_buenos, :coladas_diarias)
            """), {**fila, "programa_id": programa_id})

    print(f"\nCarga completa. programa_id={programa_id}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--archivo", default=str(
        BASE.parent / "plan de produccion mensual" / "Plan 08 Agosto.xlsx"))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    cargar(Path(args.archivo), args.dry_run)
